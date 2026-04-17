#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
香港公司审计 · 银行流水统计工具
================================

读取一个文件夹里的 DBS 星展银行（中国/香港）综合月结单 PDF，
自动解析每一笔交易，按 月 × 币种 汇总"进账"（Credit），
并产出可直接用于报价/做账的 Excel 报表。

用法:
    python3 bank_statement_analyzer.py <pdf_dir> [-o out.xlsx] [--company NAME]
                                                 [--start YYYY-MM-DD] [--end YYYY-MM-DD]
                                                 [--rates USD=7.78,EUR=8.85,...]

依赖: pdfplumber, openpyxl
    pip3 install pdfplumber openpyxl
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

try:
    import pdfplumber
except ImportError:
    sys.exit("缺少依赖 pdfplumber，请先运行: pip3 install pdfplumber openpyxl")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少依赖 openpyxl，请先运行: pip3 install pdfplumber openpyxl")


# ---------- 数据结构 ----------

@dataclass
class Txn:
    source_file: str         # 源 PDF 文件名
    account_no: str          # 账号
    currency: str            # 币种
    date: date               # 交易日期
    kind: str                # "credit" / "debit"
    amount: Decimal          # 金额（正数）
    balance: Decimal         # 交易后余额
    description: str         # 摘要（含对手方）


# ---------- DBS 综合月结单解析 ----------

# 一段账户区块的起始标志
_SECTION_RE = re.compile(
    r"FCY NRA ACCT\s+境外机构境内外汇账户\s+(\S+)\s+Currency\s+币种\s*[:：]\s*([A-Z]{3})"
)
# 亦兼容非 NRA 账户（香港本地账户）：  Product Description 里的币种信息
_SECTION_RE_GENERIC = re.compile(
    r"(?:Account No\.?\s*账号\s*[:：]?\s*)?(NRA?\d{6,}|\d{6,})\s+.{0,40}?Currency\s+币种\s*[:：]\s*([A-Z]{3})"
)

# 交易行：DD-MM-YYYY   desc...   amount   balance
_DATE_RE = re.compile(r"^(\d{2}-\d{2}-\d{4})\s+(.*)")
# 抓出每行尾部的"金额 + 余额"（可能只有余额一列，BALANCE BROUGHT FORWARD）
# 数字形如 1,234,567.89 或 0.00
_NUM = r"[\d,]+\.\d{2}"
_TRAILING_TWO_NUMS_RE = re.compile(rf"(.*?)\s+({_NUM})\s+({_NUM})\s*$")
_TRAILING_ONE_NUM_RE = re.compile(rf"(.*?)\s+({_NUM})\s*$")

_BROUGHT_FORWARD_TAG = "BALANCE BROUGHT FORWARD"
_SECTION_STOP_PHRASES = (
    "No transactions during the period",
    "此期间无交易",
)


def _to_decimal(s: str) -> Decimal:
    return Decimal(s.replace(",", ""))


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%d-%m-%Y").date()


def _split_into_account_sections(full_text: str) -> list[tuple[str, str, str]]:
    """
    把整本 PDF 的文本按「账户×币种」切分。
    返回 [(account_no, currency, body_text), ...]
    注意：同一账户跨页时，PDF 会再次出现 section header，我们在解析时再拼接。
    """
    # 切分
    positions: list[tuple[int, str, str]] = []
    for m in _SECTION_RE.finditer(full_text):
        positions.append((m.start(), m.group(1), m.group(2)))
    if not positions:
        # 兼容非 NRA 格式
        for m in _SECTION_RE_GENERIC.finditer(full_text):
            positions.append((m.start(), m.group(1), m.group(2)))

    positions.sort()
    sections: list[tuple[str, str, str]] = []
    for i, (start, acct, ccy) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(full_text)
        sections.append((acct, ccy, full_text[start:end]))
    return sections


def _parse_section(body: str) -> list[tuple[date, str, Decimal, Decimal]]:
    """
    解析一个账户区块内的交易行。
    返回 [(交易日期, 摘要, 本行金额, 交易后余额), ...]
    """
    txns: list[tuple[date, str, Decimal, Decimal]] = []
    lines = [ln.rstrip() for ln in body.splitlines()]

    # 找到首个有效锚点（承上结余）
    opening_balance: Decimal | None = None
    idx = 0
    while idx < len(lines):
        line = lines[idx]
        if _BROUGHT_FORWARD_TAG in line:
            # DD-MM-YYYY BALANCE BROUGHT FORWARD 承上结余 <balance>
            m = _TRAILING_ONE_NUM_RE.match(line)
            if m:
                try:
                    opening_balance = _to_decimal(m.group(2))
                except InvalidOperation:
                    opening_balance = None
            break
        idx += 1

    if opening_balance is None:
        # 没有开账余额就无法用 delta 法分辨方向，退化为直接读取金额
        opening_balance = Decimal("0")

    # 逐行识别交易：以 DD-MM-YYYY 开头的行即一笔交易
    # 后续无日期的行视为上一笔的摘要续行
    running: Decimal = opening_balance
    current: list[tuple[date, list[str], Decimal, Decimal]] = []
    i = idx + 1
    while i < len(lines):
        raw = lines[i].strip()
        if not raw:
            i += 1
            continue
        # 到达下一个账户块或页脚 - 交给外层逻辑
        if any(p in raw for p in _SECTION_STOP_PHRASES):
            i += 1
            continue
        # 跳过表头续行
        if raw.startswith(("Transaction Date", "交易日", "Date 日期", "Page 页数",
                           "FCY NRA ACCT", "3038496 540")):
            i += 1
            continue

        m = _DATE_RE.match(raw)
        if m:
            d = _parse_date(m.group(1))
            tail = m.group(2)
            if _BROUGHT_FORWARD_TAG in tail:
                # 跨页后又出现的承上结余，跳过
                i += 1
                continue
            # tail 里应有 金额 + 余额
            tm = _TRAILING_TWO_NUMS_RE.match(tail)
            if tm:
                desc = tm.group(1).strip()
                amount = _to_decimal(tm.group(2))
                bal = _to_decimal(tm.group(3))
                current.append((d, [desc], amount, bal))
            else:
                # 只剩一个数字 - 可能是承上结余的变体
                tm1 = _TRAILING_ONE_NUM_RE.match(tail)
                if tm1:
                    desc = tm1.group(1).strip()
                    bal = _to_decimal(tm1.group(2))
                    # 无金额列：作为调整（跳过，不计入明细）
                    i += 1
                    continue
            i += 1
            continue

        # 续行：append 到最后一笔的描述
        if current:
            current[-1][1].append(raw)
        i += 1

    result: list[tuple[date, str, Decimal, Decimal]] = []
    for d, desc_parts, amount, bal in current:
        desc = " ".join(part for part in desc_parts if part).strip()
        result.append((d, desc, amount, bal))
    return result


def parse_pdf(pdf_path: Path) -> tuple[list[Txn], str | None]:
    """
    解析单张月结单 PDF。
    返回 (交易列表, 公司名)
    """
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # 公司名（前几行里，全大写+LIMITED/LTD）
    company = None
    for line in text.splitlines()[:30]:
        s = line.strip()
        if re.fullmatch(r"[A-Z0-9 &.,()-]+(LIMITED|LTD)\.?", s):
            company = s
            break

    sections = _split_into_account_sections(text)
    all_txns: list[Txn] = []

    # 合并跨页 section：同一 (acct, ccy) 可能出现多次，按文本顺序依次拼接交易。
    grouped: dict[tuple[str, str], list[str]] = defaultdict(list)
    for acct, ccy, body in sections:
        grouped[(acct, ccy)].append(body)

    for (acct, ccy), bodies in grouped.items():
        # 取第一段里的开账余额作为 running 起点，之后各段按顺序解析
        merged_body = "\n".join(bodies)
        parsed = _parse_section(merged_body)

        # 用 delta-balance 法判断进/出
        prev_balance: Decimal | None = None
        # 再次取开账余额
        m_open = re.search(
            r"BALANCE BROUGHT FORWARD[^\n]*?(" + _NUM + r")",
            merged_body,
        )
        if m_open:
            prev_balance = _to_decimal(m_open.group(1))

        for d, desc, amount, bal in parsed:
            if prev_balance is None:
                # 无法推断方向：回退到描述判断
                kind = _guess_kind_from_desc(desc)
            else:
                delta = bal - prev_balance
                if delta > 0:
                    kind = "credit"
                elif delta < 0:
                    kind = "debit"
                else:
                    kind = _guess_kind_from_desc(desc)

            all_txns.append(
                Txn(
                    source_file=pdf_path.name,
                    account_no=acct,
                    currency=ccy,
                    date=d,
                    kind=kind,
                    amount=amount,
                    balance=bal,
                    description=desc,
                )
            )
            prev_balance = bal

    return all_txns, company


_CREDIT_KEYWORDS = ("REMITTANCE IN", "DEPOSIT", "利息存入", "CREDIT")
_DEBIT_KEYWORDS = ("REMITTANCE CHARGES", "REMITTANCE CHGS", "利息税",
                   "WITHDRAWAL", "CHARGE", "DEBIT")


def _guess_kind_from_desc(desc: str) -> str:
    u = desc.upper()
    if any(k in u for k in _CREDIT_KEYWORDS) or "利息存入" in desc:
        return "credit"
    if any(k in u for k in _DEBIT_KEYWORDS) or "利息税" in desc:
        return "debit"
    # 单独的 "REMITTANCE"（不带 IN）是汇出
    if "REMITTANCE" in u:
        return "debit"
    return "debit"


# ---------- 汇总 ----------

# 默认汇率（→HKD），来自用户模板："SGD"行 D31..H31
DEFAULT_RATES_TO_HKD: dict[str, Decimal] = {
    "HKD": Decimal("1"),
    "USD": Decimal("7.78"),
    "EUR": Decimal("8.85"),
    "JPY": Decimal("0.04987"),
    "SGD": Decimal("5.416"),
    "CNY": Decimal("1.25"),
    "GBP": Decimal("10.2"),
    "AUD": Decimal("5.1"),
}


def aggregate_credits(
    txns: list[Txn],
    start: date | None,
    end: date | None,
) -> tuple[list[str], list[tuple[int, int]], dict[tuple[int, int, str], Decimal],
           dict[tuple[int, int, str], int]]:
    """
    只统计进账。
    返回 (currencies, months, sum_map, count_map)
        months: [(年, 月), ...] 按时间升序
        sum_map[(年, 月, 币种)] = 该月该币种的进账金额合计
        count_map[(年, 月, 币种)] = 进账笔数
    """
    credits = [t for t in txns if t.kind == "credit"]
    if start:
        credits = [t for t in credits if t.date >= start]
    if end:
        credits = [t for t in credits if t.date <= end]

    currencies = sorted({t.currency for t in credits})
    months_set = {(t.date.year, t.date.month) for t in credits}
    months = sorted(months_set)

    sum_map: dict[tuple[int, int, str], Decimal] = defaultdict(lambda: Decimal("0"))
    count_map: dict[tuple[int, int, str], int] = defaultdict(int)
    for t in credits:
        k = (t.date.year, t.date.month, t.currency)
        sum_map[k] += t.amount
        count_map[k] += 1

    return currencies, months, sum_map, count_map


# ---------- Excel 生成 ----------

_THIN = Side(style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Microsoft YaHei", size=10)
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fmt_money(cell):
    cell.number_format = "#,##0.00"
    cell.font = _BODY_FONT
    cell.alignment = _RIGHT
    cell.border = _BORDER


def _header_cell(cell, text):
    cell.value = text
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER
    cell.border = _BORDER


def _body_cell(cell, text, align=None):
    cell.value = text
    cell.font = _BODY_FONT
    cell.alignment = align or _LEFT
    cell.border = _BORDER


def write_excel(
    txns: list[Txn],
    out_path: Path,
    company: str,
    period_label: str,
    rates: dict[str, Decimal],
    start: date | None,
    end: date | None,
):
    currencies, months, sum_map, count_map = aggregate_credits(txns, start, end)

    wb = Workbook()

    # ---------- Sheet 1: 进账汇总 ----------
    ws = wb.active
    ws.title = "进账汇总"

    ws["A1"] = company
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    ws["A2"] = f"账期: {period_label}"
    ws["A2"].font = Font(name="Microsoft YaHei", size=11, italic=True, color="595959")
    ws["A3"] = f"仅统计进账 (Credit) 交易；折算汇率基于下方汇率表；数据来源: {len(set(t.source_file for t in txns))} 份月结单"
    ws["A3"].font = Font(name="Microsoft YaHei", size=9, color="808080")

    header_row = 5
    # 列：月份 | 各币种金额... | 进账笔数合计 | 折合HKD合计
    headers = ["月份"] + [f"{c} 进账金额" for c in currencies] + ["进账笔数", "折合 HKD"]
    for i, h in enumerate(headers, start=1):
        _header_cell(ws.cell(row=header_row, column=i), h)

    total_by_ccy: dict[str, Decimal] = {c: Decimal("0") for c in currencies}
    total_count = 0
    total_hkd = Decimal("0")

    for r, (y, m) in enumerate(months, start=header_row + 1):
        _body_cell(ws.cell(row=r, column=1), f"{y}-{m:02d}", _CENTER)
        row_count = 0
        row_hkd = Decimal("0")
        for ci, ccy in enumerate(currencies, start=2):
            amt = sum_map.get((y, m, ccy), Decimal("0"))
            cnt = count_map.get((y, m, ccy), 0)
            cell = ws.cell(row=r, column=ci, value=float(amt) if amt else 0)
            _fmt_money(cell)
            total_by_ccy[ccy] += amt
            row_count += cnt
            rate = rates.get(ccy, Decimal("0"))
            row_hkd += amt * rate
        cell_cnt = ws.cell(row=r, column=2 + len(currencies), value=row_count)
        cell_cnt.alignment = _CENTER
        cell_cnt.font = _BODY_FONT
        cell_cnt.border = _BORDER
        total_count += row_count

        cell_hkd = ws.cell(row=r, column=3 + len(currencies), value=float(row_hkd))
        _fmt_money(cell_hkd)
        total_hkd += row_hkd

    # 合计行
    total_row = header_row + 1 + len(months)
    _body_cell(ws.cell(row=total_row, column=1), "合计", _CENTER)
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=1).font = Font(name="Microsoft YaHei", size=10, bold=True)
    for ci, ccy in enumerate(currencies, start=2):
        c = ws.cell(row=total_row, column=ci, value=float(total_by_ccy[ccy]))
        _fmt_money(c)
        c.fill = _TOTAL_FILL
        c.font = Font(name="Microsoft YaHei", size=10, bold=True)
    c = ws.cell(row=total_row, column=2 + len(currencies), value=total_count)
    c.alignment = _CENTER
    c.fill = _TOTAL_FILL
    c.border = _BORDER
    c.font = Font(name="Microsoft YaHei", size=10, bold=True)
    c = ws.cell(row=total_row, column=3 + len(currencies), value=float(total_hkd))
    _fmt_money(c)
    c.fill = _TOTAL_FILL
    c.font = Font(name="Microsoft YaHei", size=10, bold=True)

    # 汇率表
    rate_row = total_row + 2
    ws.cell(row=rate_row, column=1, value="汇率→HKD").font = Font(
        name="Microsoft YaHei", size=10, bold=True, color="595959"
    )
    for ci, ccy in enumerate(currencies, start=2):
        c = ws.cell(row=rate_row, column=ci, value=float(rates.get(ccy, Decimal("0"))))
        c.number_format = "0.0000"
        c.font = Font(name="Microsoft YaHei", size=10, italic=True, color="595959")
        c.alignment = _RIGHT

    # 列宽
    ws.column_dimensions["A"].width = 12
    for ci in range(2, 2 + len(currencies)):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.column_dimensions[get_column_letter(2 + len(currencies))].width = 12
    ws.column_dimensions[get_column_letter(3 + len(currencies))].width = 16

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # ---------- Sheet 2: 进账明细 ----------
    ws2 = wb.create_sheet("进账明细")
    headers2 = ["日期", "月份", "账号", "币种", "金额", "交易后余额", "摘要/对手方", "源文件"]
    for i, h in enumerate(headers2, start=1):
        _header_cell(ws2.cell(row=1, column=i), h)

    credits = sorted(
        [t for t in txns if t.kind == "credit"
         and (not start or t.date >= start)
         and (not end or t.date <= end)],
        key=lambda t: (t.date, t.currency, t.account_no),
    )
    for r, t in enumerate(credits, start=2):
        _body_cell(ws2.cell(row=r, column=1), t.date.strftime("%Y-%m-%d"), _CENTER)
        _body_cell(ws2.cell(row=r, column=2), f"{t.date.year}-{t.date.month:02d}", _CENTER)
        _body_cell(ws2.cell(row=r, column=3), t.account_no, _CENTER)
        _body_cell(ws2.cell(row=r, column=4), t.currency, _CENTER)
        _fmt_money(ws2.cell(row=r, column=5, value=float(t.amount)))
        _fmt_money(ws2.cell(row=r, column=6, value=float(t.balance)))
        _body_cell(ws2.cell(row=r, column=7), t.description, _LEFT)
        _body_cell(ws2.cell(row=r, column=8), t.source_file, _CENTER)

    widths = [12, 10, 22, 8, 16, 16, 60, 20]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---------- Sheet 3: 全部交易（含支出，供核对） ----------
    ws3 = wb.create_sheet("全部交易")
    headers3 = ["日期", "账号", "币种", "方向", "金额", "交易后余额", "摘要", "源文件"]
    for i, h in enumerate(headers3, start=1):
        _header_cell(ws3.cell(row=1, column=i), h)

    all_sorted = sorted(
        [t for t in txns
         if (not start or t.date >= start) and (not end or t.date <= end)],
        key=lambda t: (t.date, t.currency, t.account_no),
    )
    for r, t in enumerate(all_sorted, start=2):
        _body_cell(ws3.cell(row=r, column=1), t.date.strftime("%Y-%m-%d"), _CENTER)
        _body_cell(ws3.cell(row=r, column=2), t.account_no, _CENTER)
        _body_cell(ws3.cell(row=r, column=3), t.currency, _CENTER)
        kind_cell = ws3.cell(row=r, column=4, value="进账" if t.kind == "credit" else "支出")
        kind_cell.alignment = _CENTER
        kind_cell.font = Font(
            name="Microsoft YaHei", size=10, bold=True,
            color="C00000" if t.kind == "credit" else "595959",
        )
        kind_cell.border = _BORDER
        _fmt_money(ws3.cell(row=r, column=5, value=float(t.amount)))
        _fmt_money(ws3.cell(row=r, column=6, value=float(t.balance)))
        _body_cell(ws3.cell(row=r, column=7), t.description, _LEFT)
        _body_cell(ws3.cell(row=r, column=8), t.source_file, _CENTER)

    for i, w in enumerate([12, 22, 8, 8, 16, 16, 60, 20], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ---------- Sheet 4: 源文件索引 ----------
    ws4 = wb.create_sheet("源文件")
    _header_cell(ws4.cell(row=1, column=1), "源文件")
    _header_cell(ws4.cell(row=1, column=2), "覆盖月份")
    _header_cell(ws4.cell(row=1, column=3), "进账笔数")
    _header_cell(ws4.cell(row=1, column=4), "出账笔数")

    by_file: dict[str, list[Txn]] = defaultdict(list)
    for t in txns:
        by_file[t.source_file].append(t)

    for r, fn in enumerate(sorted(by_file.keys()), start=2):
        items = by_file[fn]
        months_in = sorted({f"{t.date.year}-{t.date.month:02d}" for t in items})
        cred = sum(1 for t in items if t.kind == "credit")
        deb = sum(1 for t in items if t.kind == "debit")
        _body_cell(ws4.cell(row=r, column=1), fn, _LEFT)
        _body_cell(ws4.cell(row=r, column=2), ", ".join(months_in), _LEFT)
        _body_cell(ws4.cell(row=r, column=3), cred, _CENTER)
        _body_cell(ws4.cell(row=r, column=4), deb, _CENTER)

    for i, w in enumerate([42, 24, 12, 12], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- CLI ----------

def _collect_pdfs(input_dir: Path) -> list[Path]:
    pdfs = []
    for p in input_dir.rglob("*.pdf"):
        if p.name.startswith("._") or p.name.startswith("~$"):
            continue
        pdfs.append(p)
    return sorted(pdfs)


def _parse_rates(s: str) -> dict[str, Decimal]:
    out = dict(DEFAULT_RATES_TO_HKD)
    if not s:
        return out
    for pair in s.split(","):
        if "=" not in pair:
            continue
        k, v = pair.split("=", 1)
        try:
            out[k.strip().upper()] = Decimal(v.strip())
        except InvalidOperation:
            pass
    return out


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="解析 DBS 月结单，生成按月×币种汇总的 Excel 报表",
    )
    parser.add_argument("input_dir", help="包含 PDF 月结单的文件夹（会递归查找）")
    parser.add_argument("-o", "--output", default=None, help="输出 xlsx 路径")
    parser.add_argument("--company", default=None, help="公司名（缺省自动识别）")
    parser.add_argument("--start", default=None, help="审计期起 YYYY-MM-DD")
    parser.add_argument("--end", default=None, help="审计期止 YYYY-MM-DD")
    parser.add_argument(
        "--rates",
        default=None,
        help="汇率表 (CCY=rate,...)，例: USD=7.78,EUR=8.85。默认值已内置。",
    )
    args = parser.parse_args(argv)

    input_dir = Path(args.input_dir).expanduser().resolve()
    if not input_dir.is_dir():
        sys.exit(f"输入目录不存在: {input_dir}")

    pdfs = _collect_pdfs(input_dir)
    if not pdfs:
        sys.exit(f"在 {input_dir} 下未找到 PDF 月结单")

    print(f"找到 {len(pdfs)} 份 PDF，开始解析…")

    all_txns: list[Txn] = []
    detected_company: str | None = None
    for p in pdfs:
        try:
            txns, company = parse_pdf(p)
        except Exception as e:
            print(f"  !! {p.name} 解析失败: {e}", file=sys.stderr)
            continue
        all_txns.extend(txns)
        if company and not detected_company:
            detected_company = company
        cred = sum(1 for t in txns if t.kind == "credit")
        print(f"  • {p.name}: {len(txns)} 笔交易 (进账 {cred})")

    if not all_txns:
        sys.exit("未解析到任何交易，请检查 PDF 是否为 DBS 综合月结单。")

    company = args.company or detected_company or "（未识别公司名）"
    start = datetime.strptime(args.start, "%Y-%m-%d").date() if args.start else None
    end = datetime.strptime(args.end, "%Y-%m-%d").date() if args.end else None

    all_dates = sorted(t.date for t in all_txns)
    period_label = (
        f"{(start or all_dates[0]).strftime('%Y.%-m.%-d')}"
        f" - {(end or all_dates[-1]).strftime('%Y.%-m.%-d')}"
    )

    rates = _parse_rates(args.rates or "")

    safe_company = re.sub(r"[\\/:*?\"<>|]+", "", company).strip()
    default_name = f"进账流水统计-{safe_company}.xlsx"
    out_path = Path(args.output).expanduser().resolve() if args.output \
        else input_dir.parent / default_name

    write_excel(all_txns, out_path, company, period_label, rates, start, end)

    total_credits = sum(1 for t in all_txns if t.kind == "credit")
    total_debits = sum(1 for t in all_txns if t.kind == "debit")
    print(
        f"\n✓ 完成: {out_path}\n"
        f"  公司: {company}\n"
        f"  账期: {period_label}\n"
        f"  交易合计: {len(all_txns)} 笔 (进账 {total_credits} / 支出 {total_debits})"
    )


if __name__ == "__main__":
    main()
